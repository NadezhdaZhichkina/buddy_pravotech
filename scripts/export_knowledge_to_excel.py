#!/usr/bin/env python3
"""
Экспорт базы знаний в Excel.
  python scripts/export_knowledge_to_excel.py

Создаёт файл knowledge_base.xlsx в папке buddy.
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# Загружаем полную базу из seed_knowledge.py
import importlib.util
seed_path = Path(__file__).parent / "seed_knowledge.py"
spec = importlib.util.spec_from_file_location("seed_knowledge", seed_path)
seed_mod = importlib.util.module_from_spec(spec)
spec.loader.exec_module(seed_mod)
items = seed_mod.SEED_ITEMS

wb = Workbook()
ws = wb.active
ws.title = "База знаний"

# Заголовки
headers = ["№", "Вопрос", "Ответ", "Теги"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(wrap_text=True, vertical="top")

# Данные
for i, item in enumerate(items, 1):
    ws.cell(row=i + 1, column=1, value=i)
    ws.cell(row=i + 1, column=2, value=item["question"])
    ws.cell(row=i + 1, column=3, value=item["answer"])
    ws.cell(row=i + 1, column=4, value=item.get("tags", ""))

# Ширина колонок
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 80
ws.column_dimensions["D"].width = 30

# Перенос текста для ответов
for row in range(2, len(items) + 2):
    for col in range(1, 5):
        ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

out_path = Path(__file__).resolve().parent.parent / "knowledge_base.xlsx"
wb.save(out_path)
print(f"Создан файл: {out_path}")
print(f"Записей: {len(items)}")
